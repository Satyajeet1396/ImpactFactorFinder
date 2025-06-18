import streamlit as st
import pandas as pd
from rapidfuzz import process, fuzz
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
from functools import lru_cache
from tqdm import tqdm
import qrcode
import base64
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np

# Reuse the existing journal standardization functions
JOURNAL_REPLACEMENTS = {
    'intl': 'international',
    'int': 'international',
    'natl': 'national',
    'nat': 'national',
    'sci': 'science',
    'med': 'medical',
    'res': 'research',
    'tech': 'technology',
    'eng': 'engineering',
    'phys': 'physics',
    'chem': 'chemistry',
    'bio': 'biology',
    'env': 'environmental',
    'mgmt': 'management',
    'dev': 'development',
    'edu': 'education',
    'univ': 'university',
    'j\\.': 'journal',
    'jr\\.': 'journal',
    'jrnl\\.': 'journal',
    'proc\\.': 'proceedings',
    'rev\\.': 'review',
    'q\\.': 'quarterly',
}

@lru_cache(maxsize=10000)
def standardize_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower().strip()
    text = text.replace('&', 'and')
    text = re.sub(r'[-:]', ' ', text)
    text = re.sub(r'\([^)]*?(print|online|www|web|issn|doi).*?\)', '', text, flags=re.IGNORECASE)
    for abbr, full in JOURNAL_REPLACEMENTS.items():
        text = re.sub(rf'\b{abbr}\b', full, text, flags=re.IGNORECASE)
    text = re.sub(r'[^\w\s]', ' ', text)
    return ' '.join(text.split())

def read_file(file):
    file_extension = file.name.split('.')[-1].lower()
    if file_extension == 'xlsx':
        return pd.read_excel(file)
    elif file_extension == 'csv':
        return pd.read_csv(file)
    else:
        raise ValueError("Unsupported file format. Please upload either CSV or XLSX file.")

def create_statistics_visualizations(results_df, ref_df):
    """Create comprehensive statistics and visualizations"""
    
    # Extract numeric JIF values for matched journals
    matched_journals = results_df[results_df['Match Score'] > 0].copy()
    
    # Parse JIF values (handle multiple values separated by commas)
    jif_values = []
    quartiles = []
    
    for idx, row in matched_journals.iterrows():
        if pd.notna(row['JIF 2024']) and row['JIF 2024'] != '':
            try:
                # Handle multiple JIF values (take the first one)
                jif_str = str(row['JIF 2024']).split(',')[0].strip()
                jif_val = float(jif_str)
                jif_values.append(jif_val)
                quartiles.append(row['JIF Quartile'])
            except:
                continue
    
    # Create visualizations
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('JIF Distribution', 'Quartile Distribution', 
                       'JIF by Quartile', 'Match Score Distribution'),
        specs=[[{"type": "histogram"}, {"type": "pie"}],
               [{"type": "box"}, {"type": "histogram"}]]
    )
    
    # JIF Distribution
    if jif_values:
        fig.add_trace(go.Histogram(x=jif_values, name="JIF Distribution", 
                                  nbinsx=20, marker_color='lightblue'), 
                     row=1, col=1)
    
    # Quartile Distribution
    if quartiles:
        quartile_counts = pd.Series(quartiles).value_counts()
        fig.add_trace(go.Pie(labels=quartile_counts.index, values=quartile_counts.values,
                            name="Quartiles"), row=1, col=2)
    
    # JIF by Quartile Box Plot
    if jif_values and quartiles:
        quartile_df = pd.DataFrame({'JIF': jif_values, 'Quartile': quartiles})
        for q in ['Q1', 'Q2', 'Q3', 'Q4']:
            q_data = quartile_df[quartile_df['Quartile'] == q]['JIF']
            if len(q_data) > 0:
                fig.add_trace(go.Box(y=q_data, name=q, boxpoints='outliers'), 
                             row=2, col=1)
    
    # Match Score Distribution
    match_scores = results_df['Match Score']
    fig.add_trace(go.Histogram(x=match_scores, name="Match Scores", 
                              nbinsx=10, marker_color='lightgreen'), 
                 row=2, col=2)
    
    fig.update_layout(height=800, showlegend=False, 
                     title_text="Journal Analysis Dashboard")
    
    return fig

def calculate_comprehensive_statistics(results_df, ref_df):
    """Calculate comprehensive statistics about the processed data"""
    
    total_journals = len(results_df)
    matched_journals = results_df[results_df['Match Score'] > 0]
    perfect_matches = results_df[results_df['Match Score'] == 100]
    good_matches = results_df[(results_df['Match Score'] >= 90) & (results_df['Match Score'] < 100)]
    no_matches = results_df[results_df['Match Score'] == 0]
    
    # JIF Statistics
    jif_values = []
    quartile_counts = {'Q1': 0, 'Q2': 0, 'Q3': 0, 'Q4': 0}
    
    for idx, row in matched_journals.iterrows():
        if pd.notna(row['JIF 2024']) and row['JIF 2024'] != '':
            try:
                jif_str = str(row['JIF 2024']).split(',')[0].strip()
                jif_val = float(jif_str)
                jif_values.append(jif_val)
                
                # Count quartiles
                if pd.notna(row['JIF Quartile']):
                    q = row['JIF Quartile']
                    if q in quartile_counts:
                        quartile_counts[q] += 1
            except:
                continue
    
    stats = {
        'total_journals': total_journals,
        'matched_journals': len(matched_journals),
        'perfect_matches': len(perfect_matches),
        'good_matches': len(good_matches),
        'no_matches': len(no_matches),
        'match_rate': len(matched_journals) / total_journals * 100 if total_journals > 0 else 0,
        'perfect_match_rate': len(perfect_matches) / total_journals * 100 if total_journals > 0 else 0,
        'jif_count': len(jif_values),
        'jif_stats': {
            'mean': np.mean(jif_values) if jif_values else 0,
            'median': np.median(jif_values) if jif_values else 0,
            'std': np.std(jif_values) if jif_values else 0,
            'min': np.min(jif_values) if jif_values else 0,
            'max': np.max(jif_values) if jif_values else 0,
            'q25': np.percentile(jif_values, 25) if jif_values else 0,
            'q75': np.percentile(jif_values, 75) if jif_values else 0
        },
        'quartile_distribution': quartile_counts,
        'high_impact_journals': len([x for x in jif_values if x >= 10]) if jif_values else 0,
        'top_tier_journals': len([x for x in jif_values if x >= 20]) if jif_values else 0,
        'elite_journals': len([x for x in jif_values if x >= 50]) if jif_values else 0
    }
    
    return stats

def process_single_file(user_df, ref_df):
    # Find the "Source title" column
    source_title_col = None
    for col in user_df.columns:
        if 'source title' in str(col).lower():
            source_title_col = col
            break
    
    if source_title_col is None:
        st.error("No 'Source title' column found in the input file. Please ensure your file has a column containing 'Source title'.")
        return None
    
    # Create a copy of journal names for processing
    journals = user_df[source_title_col].astype(str).apply(standardize_text)
    
    # Create reference dictionary with JIF and Quartile information
    ref_dict = {}
    
    # Assuming reference file has columns: Journal Name, JIF 2024, JIF Quartile
    ref_df_clean = ref_df.copy()
    
    # Standardize reference journal names
    ref_df_clean.iloc[:, 1] = ref_df_clean.iloc[:, 1].astype(str).apply(standardize_text)  # Journal Name column
    
    for idx, row in ref_df_clean.iterrows():
        journal = row.iloc[1]  # Journal Name
        jif = row.iloc[2] if len(row) > 2 else ''  # JIF 2024
        quartile = row.iloc[3] if len(row) > 3 else ''  # JIF Quartile
        
        if journal not in ref_dict:
            ref_dict[journal] = []
        ref_dict[journal].append({
            'jif': jif,
            'quartile': quartile
        })
    
    ref_journals = ref_df_clean.iloc[:, 1].tolist()
    journal_list = journals.tolist()
    
    results = []
    for journal in tqdm(journal_list, desc="Processing journals"):
        if pd.isna(journal) or str(journal).strip() == "":
            results.append(("No journal name", "No match found", 0, "", ""))
            continue
            
        if journal in ref_dict:
            # Perfect match found
            match_info = ref_dict[journal][0]  # Take first match
            results.append((journal, journal, 100, match_info['jif'], match_info['quartile']))
            continue
            
        # Fuzzy matching
        match = process.extractOne(journal, ref_journals, scorer=fuzz.ratio, score_cutoff=90)
        if match:
            match_info = ref_dict[match[0]][0]  # Take first match
            results.append((journal, match[0], match[1], match_info['jif'], match_info['quartile']))
        else:
            results.append((journal, "No match found", 0, "", ""))
    
    # Create DataFrame with match results
    new_columns = ['Processed Journal Name', 'Best Match', 'Match Score', 'JIF 2024', 'JIF Quartile']
    results_df = pd.DataFrame(results, columns=new_columns)
    
    # Calculate and display statistics
    stats = calculate_comprehensive_statistics(results_df, ref_df)
    
    # Display comprehensive statistics
    st.write("### 📊 Comprehensive Analysis")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Journals", stats['total_journals'])
        st.metric("Match Rate", f"{stats['match_rate']:.1f}%")
    with col2:
        st.metric("Perfect Matches", stats['perfect_matches'])
        st.metric("Good Matches", stats['good_matches'])
    with col3:
        st.metric("No Matches", stats['no_matches'])
        st.metric("JIF Available", stats['jif_count'])
    with col4:
        st.metric("High Impact (≥10)", stats['high_impact_journals'])
        st.metric("Elite Journals (≥50)", stats['elite_journals'])
    
    # JIF Statistics
    if stats['jif_count'] > 0:
        st.write("### 📈 Journal Impact Factor Statistics")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("**Descriptive Statistics:**")
            jif_stats = stats['jif_stats']
            st.write(f"- Mean JIF: {jif_stats['mean']:.2f}")
            st.write(f"- Median JIF: {jif_stats['median']:.2f}")
            st.write(f"- Standard Deviation: {jif_stats['std']:.2f}")
            st.write(f"- Range: {jif_stats['min']:.2f} - {jif_stats['max']:.2f}")
        
        with col2:
            st.write("**Quartile Distribution:**")
            for q, count in stats['quartile_distribution'].items():
                percentage = count / stats['jif_count'] * 100 if stats['jif_count'] > 0 else 0
                st.write(f"- {q}: {count} journals ({percentage:.1f}%)")
    
    # Quality Assessment
    st.write("### 🎯 Quality Assessment")
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**Journal Quality Tiers:**")
        st.write(f"- Elite (JIF ≥ 50): {stats['elite_journals']} journals")
        st.write(f"- Top Tier (JIF ≥ 20): {stats['top_tier_journals']} journals")
        st.write(f"- High Impact (JIF ≥ 10): {stats['high_impact_journals']} journals")
        st.write(f"- Total with JIF: {stats['jif_count']} journals")
    
    with col2:
        st.write("**Matching Quality:**")
        st.write(f"- Perfect matches: {stats['perfect_matches']} ({stats['perfect_match_rate']:.1f}%)")
        st.write(f"- Good matches: {stats['good_matches']}")
        st.write(f"- Overall match rate: {stats['match_rate']:.1f}%")
        st.write(f"- Data completeness: {(stats['jif_count']/stats['matched_journals']*100) if stats['matched_journals'] > 0 else 0:.1f}%")
    
    # Create and display visualizations
    if stats['jif_count'] > 0:
        try:
            fig = create_statistics_visualizations(results_df, ref_df)
            st.plotly_chart(fig, use_container_width=True)
        except Exception as e:
            st.warning(f"Could not create visualizations: {str(e)}")
    
    # Add processed journal name and match results
    final_df = pd.concat([
        user_df,
        results_df
    ], axis=1)
    
    # Sort by Match Score in ascending order (poorest matches first)
    final_df = final_df.sort_values(by='Match Score', ascending=True)
    
    # Store the new column names for highlighting in Excel
    final_df.attrs['new_columns'] = new_columns
    
    return final_df

def save_results(df, file_format='xlsx'):
    output = BytesIO()
    
    if file_format == 'xlsx':
        # Save to Excel with styled headers
        writer = pd.ExcelWriter(output, engine='openpyxl')
        df.to_excel(writer, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Create fill styles for different columns
        header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        jif_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
        quartile_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        
        # Get new column names from DataFrame attributes
        new_columns = df.attrs.get('new_columns', [])
        
        # Apply highlighting to headers
        for cell in worksheet[1]:
            if cell.value in new_columns:
                if 'JIF 2024' in str(cell.value):
                    cell.fill = jif_fill
                elif 'Quartile' in str(cell.value):
                    cell.fill = quartile_fill
                else:
                    cell.fill = header_fill
                cell.font = Font(color='FFFFFF', bold=True)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        writer.close()
    else:
        # For CSV, just save normally
        df.to_csv(output, index=False)
    
    output.seek(0)
    return output

# Streamlit app
st.set_page_config(page_title="Enhanced Journal Impact Factor Processor", layout="wide")
st.title("🔬 Enhanced Journal Impact Factor Processor")
st.markdown("*Find JIF values, quartiles, and comprehensive analytics for your journal lists*")

# Add collapsible app information
with st.expander("ℹ️ Click here to learn about this enhanced app", expanded=False):
    st.markdown("""
        <style>
        .app-info {
            padding: 20px;
            border-radius: 10px;
            background-color: #f0f2f6;
            margin: 10px 0;
        }
        .app-info h3 {
            color: #0066cc;
            margin-top: 20px;
            margin-bottom: 10px;
        }
        .app-info ul, .app-info ol {
            margin-bottom: 20px;
        }
        </style>
        <div class="app-info">
        <h3>🚀 New Features</h3>
        <ul>
        <li><strong>JIF 2024 Values:</strong> Get the latest Journal Impact Factor data</li>
        <li><strong>JIF Quartiles:</strong> Understand journal ranking (Q1, Q2, Q3, Q4)</li>
        <li><strong>Comprehensive Statistics:</strong> Detailed analytics and visualizations</li>
        <li><strong>Quality Assessment:</strong> Categorize journals by impact level</li>
        <li><strong>Interactive Dashboard:</strong> Visual analysis of your journal portfolio</li>
        </ul>
        
        <h3>📊 What You Get</h3>
        <ul>
        <li>Match statistics and success rates</li>
        <li>JIF distribution analysis</li>
        <li>Quartile breakdown of your journals</li>
        <li>Quality tier classification (Elite, Top Tier, High Impact)</li>
        <li>Interactive charts and visualizations</li>
        <li>Enhanced Excel output with color coding</li>
        </ul>
        
        <h3>🎯 Journal Quality Tiers</h3>
        <ul>
        <li><strong>Elite Journals:</strong> JIF ≥ 50 (Top-tier prestigious journals)</li>
        <li><strong>Top Tier:</strong> JIF ≥ 20 (Highly respected journals)</li>
        <li><strong>High Impact:</strong> JIF ≥ 10 (Well-regarded journals)</li>
        <li><strong>Q1 Quartile:</strong> Top 25% in their field</li>
        <li><strong>Q2-Q4 Quartiles:</strong> Remaining journals by impact ranking</li>
        </ul>
        </div>
    """, unsafe_allow_html=True)

st.write("Upload multiple journal lists to process them with enhanced JIF analysis.")

# Initialize session states
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = set()
if 'processed_results' not in st.session_state:
    st.session_state.processed_results = {}

# File uploads
uploaded_files = st.file_uploader("Upload Your Journal Lists (Excel/CSV)", type=["xlsx", "csv"], accept_multiple_files=True, key="file_uploader")

# Reference file loading with error handling
try:
    # You should replace this with your actual reference file path/URL
    reference_file_url = "https://github.com/Satyajeet1396/ImpactFactorFinder/blob/c48c0176e314f05e25f7a7c219f67ab9ab6d5237/IF2025JCR.xlsx"
    ref_df = pd.read_excel(reference_file_url)
    
    # Display reference file info
    st.success(f"✅ Successfully loaded reference database with {len(ref_df)} journal entries")
    
    # Show sample of reference data
    with st.expander("📋 View Reference Database Sample", expanded=False):
        st.dataframe(ref_df.head(10))
        st.write(f"**Columns:** {list(ref_df.columns)}")
        
except Exception as e:
    st.error(f"❌ Error loading reference database: {str(e)}")
    st.info("Please ensure your reference file has columns: Rank, Journal Name, JIF 2024, JIF Quartile")
    st.stop()

if uploaded_files:
    # Process each file that hasn't been processed yet
    for uploaded_file in uploaded_files:
        file_identifier = f"{uploaded_file.name}_{uploaded_file.size}"
        
        if file_identifier not in st.session_state.processed_files:
            st.write(f"🔄 Processing {uploaded_file.name}...")
            
            try:
                # Read and process the file
                user_df = read_file(uploaded_file)
                st.info(f"📄 Found {len(user_df)} entries in {uploaded_file.name}")
                
                with st.spinner(f"Processing {uploaded_file.name}..."):
                    results_df = process_single_file(user_df, ref_df)
                    
                    if results_df is not None:
                        # Save results to session state
                        output_format = uploaded_file.name.split('.')[-1].lower()
                        output_file = save_results(results_df, output_format)
                        
                        st.session_state.processed_results[file_identifier] = {
                            'results_df': results_df,
                            'output_file': output_file,
                            'output_format': output_format,
                            'filename': uploaded_file.name
                        }
                        
                        # Mark file as processed
                        st.session_state.processed_files.add(file_identifier)
                        st.success(f"✅ Successfully processed {uploaded_file.name}")
            
            except Exception as e:
                st.error(f"❌ Error processing {uploaded_file.name}: {str(e)}")
                continue
    
    # Display results for all processed files
    if st.session_state.processed_results:
        st.write("## 📋 Processed Files Results")
        for file_id, data in st.session_state.processed_results.items():
            with st.expander(f"📊 Results for {data['filename']}", expanded=True):
                # Create download button for this file
                output_filename = f"{data['filename'].rsplit('.', 1)[0]}_enhanced_matched.{data['output_format']}"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if data['output_format'] == 'xlsx' else "text/csv"
                
                st.download_button(
                    label=f"📥 Download Enhanced Results for {data['filename']}",
                    data=data['output_file'],
                    file_name=output_filename,
                    mime=mime_type,
                    key=f"download_{file_id}",
                    type="primary"
                )
                
                # Show sample results with key columns
                st.write("**Sample Results with JIF Data:**")
                display_cols = ['Processed Journal Name', 'Best Match', 'Match Score', 'JIF 2024', 'JIF Quartile']
                available_cols = [col for col in display_cols if col in data['results_df'].columns]
                st.dataframe(data['results_df'][available_cols].head(10))
    
    # Add a button to clear processed files and start fresh
    if st.button("🔄 Clear All and Process New Files", type="secondary"):
        st.session_state.processed_files.clear()
        st.session_state.processed_results.clear()
        st.rerun()
            
else:
    st.info("📤 Please upload one or more journal lists (XLSX or CSV format) to get started.")
    st.markdown("""
    **Expected Format:**
    - Your file should contain a column with 'Source title' or similar
    - The app will automatically detect and process journal names
    - Results will include JIF 2024 values and quartile rankings
    """)

st.divider()

# Footer and support section
col1, col2 = st.columns([2, 1])
with col1:
    st.info("Created by Dr. Satyajeet Patil")
    st.info("For more research tools visit: https://patilsatyajeet.wixsite.com/home/python")
with col2:
    st.metric("Enhanced Features", "5+", delta="New in this version")

# Support section in expander
with st.expander("🤝 Support Our Research", expanded=False):
    st.markdown("""
        <div style='text-align: center; padding: 1rem; background-color: #f0f2f6; border-radius: 10px; margin: 1rem 0;'>
            <h3>🙏 Your Support Makes a Difference!</h3>
            <p>Your contribution helps us continue developing free tools for the research community.</p>
            <p>Every donation, no matter how small, fuels our research journey!</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Two columns for QR code and Buy Me a Coffee button
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### UPI Payment")
        # Generate UPI QR code
        upi_url = "upi://pay?pa=satyajeet1396@oksbi&pn=Satyajeet Patil&cu=INR"
        qr = qrcode.make(upi_url)
        
        # Save QR code to BytesIO
        buffer = BytesIO()
        qr.save(buffer, format="PNG")
        buffer.seek(0)
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        # Display QR code with message
        st.markdown("Scan to pay: **satyajeet1396@oksbi**")
        st.markdown(
            f"""
            <div style="display: flex; justify-content: center; align-items: center;">
                <img src="data:image/png;base64,{qr_base64}" width="200">
            </div>
            """,
            unsafe_allow_html=True
        )
    
    with col2:
        st.markdown("#### Buy Me a Coffee")
        st.markdown("Support through Buy Me a Coffee platform:")
        # Buy Me a Coffee button
        st.markdown(
            """
            <div style="display: flex; justify-content: center; align-items: center; height: 100%;">
                <a href="https://www.buymeacoffee.com/researcher13" target="_blank">
                    <img src="https://img.buymeacoffee.com/button-api/?text=Support our Research&emoji=&slug=researcher13&button_colour=FFDD00&font_colour=000000&font_family=Cookie&outline_colour=000000&coffee_colour=ffffff" alt="Support our Research"/>
                </a>
            </div>
            """,
            unsafe_allow_html=True
        )

st.info("🚀 A small donation from you can fuel our research journey, turning ideas into breakthroughs that can change lives!")
